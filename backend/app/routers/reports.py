from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
from ..database import get_db
from ..models.inspection import Inspection, InspectionStatus
from ..models.corrective_action import CorrectiveAction, ActionStatus
from ..models.certification import TeamCertification, CertStatus
from ..models.clinic import Clinic
from ..models.user import User
from .deps import get_current_user, require_admin_or_auditor

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from ..models.user import UserRole
    now = datetime.utcnow()

    q_insp = db.query(Inspection)
    q_actions = db.query(CorrectiveAction)
    q_certs = db.query(TeamCertification)

    if current_user.role == UserRole.manager:
        managed_ids = db.query(Clinic.id).filter(Clinic.manager_id == current_user.id).subquery()
        q_insp = q_insp.filter(Inspection.clinic_id.in_(managed_ids))
        q_actions = q_actions.filter(CorrectiveAction.clinic_id.in_(managed_ids))

    total_inspections = q_insp.count()
    approved = q_insp.filter(Inspection.status == InspectionStatus.approved).count()
    pending_review = q_insp.filter(Inspection.status.in_([InspectionStatus.submitted, InspectionStatus.under_review])).count()
    avg_score = db.query(func.avg(Inspection.compliance_score)).filter(Inspection.compliance_score.isnot(None)).scalar()

    open_actions = q_actions.filter(CorrectiveAction.status.in_([ActionStatus.open, ActionStatus.in_progress, ActionStatus.pending_verification])).count()
    overdue_actions = q_actions.filter(
        CorrectiveAction.status.in_([ActionStatus.open, ActionStatus.in_progress]),
        CorrectiveAction.due_date < now.date(),
    ).count()

    total_certs = q_certs.count()
    completed_certs = q_certs.filter(TeamCertification.status == CertStatus.completed).count()
    expiring_certs = q_certs.filter(
        TeamCertification.status == CertStatus.completed,
        TeamCertification.expires_at.isnot(None),
        TeamCertification.expires_at <= now + timedelta(days=30),
        TeamCertification.expires_at > now,
    ).count()

    # Clinics by risk level
    clinics = db.query(Clinic).filter(Clinic.is_active == True).all()
    risk_breakdown = {"low": 0, "medium": 0, "high": 0, "critical": 0, "unknown": 0}
    clinic_scores = []
    for c in clinics:
        latest = (db.query(Inspection)
                  .filter(Inspection.clinic_id == c.id, Inspection.compliance_score.isnot(None))
                  .order_by(Inspection.submitted_at.desc()).first())
        if latest:
            risk = latest.risk_level or "unknown"
            risk_breakdown[risk] = risk_breakdown.get(risk, 0) + 1
            clinic_scores.append({
                "clinic_id": c.id, "clinic_name": c.name,
                "score": latest.compliance_score, "risk_level": latest.risk_level,
                "last_inspection": str(latest.submitted_at) if latest.submitted_at else None,
            })
        else:
            risk_breakdown["unknown"] += 1

    recent = (db.query(Inspection)
              .filter(Inspection.compliance_score.isnot(None))
              .order_by(Inspection.submitted_at.desc()).limit(10).all())

    trend = []
    for i in range(5, -1, -1):
        start = (now - timedelta(days=30 * (i + 1))).replace(day=1)
        end = (now - timedelta(days=30 * i)).replace(day=1)
        avg = db.query(func.avg(Inspection.compliance_score)).filter(
            Inspection.submitted_at >= start, Inspection.submitted_at < end,
            Inspection.compliance_score.isnot(None)).scalar()
        trend.append({"month": start.strftime("%b %Y"), "avg_score": round(avg or 0, 1)})

    return {
        "summary": {
            "total_inspections": total_inspections,
            "approved_inspections": approved,
            "pending_review": pending_review,
            "avg_compliance_score": round(avg_score or 0, 1),
            "open_corrective_actions": open_actions,
            "overdue_corrective_actions": overdue_actions,
            "total_certifications": total_certs,
            "completed_certifications": completed_certs,
            "expiring_certifications": expiring_certs,
        },
        "risk_breakdown": risk_breakdown,
        "clinic_scores": clinic_scores,
        "trend": trend,
        "recent_inspections": [
            {
                "id": i.id,
                "clinic_name": i.clinic.name if i.clinic else None,
                "score": i.compliance_score,
                "risk_level": i.risk_level,
                "status": i.status.value if i.status else None,
                "submitted_at": str(i.submitted_at) if i.submitted_at else None,
            }
            for i in recent
        ],
    }


@router.get("/my-tasks")
def my_tasks(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Staff-level view: my assigned tasks, my certifications, my inspection history."""
    actions = db.query(CorrectiveAction).filter(
        CorrectiveAction.assigned_to == current_user.id,
        CorrectiveAction.status.in_([ActionStatus.open, ActionStatus.in_progress]),
    ).order_by(CorrectiveAction.due_date).limit(20).all()

    certs = db.query(TeamCertification).filter(
        TeamCertification.participant_email == current_user.email,
    ).order_by(TeamCertification.created_at.desc()).limit(20).all()

    inspections = db.query(Inspection).filter(
        Inspection.inspector_id == current_user.id,
    ).order_by(Inspection.created_at.desc()).limit(10).all()

    return {
        "assigned_actions": [
            {"id": a.id, "title": a.title, "status": a.status.value,
             "priority": a.priority, "due_date": str(a.due_date) if a.due_date else None,
             "clinic_name": a.clinic.name if a.clinic else None}
            for a in actions
        ],
        "certifications": [
            {"id": c.id, "course_title": c.course.title if c.course else None,
             "status": c.status.value, "score": c.score,
             "expires_at": str(c.expires_at) if c.expires_at else None}
            for c in certs
        ],
        "recent_inspections": [
            {"id": i.id, "clinic_name": i.clinic.name if i.clinic else None,
             "score": i.compliance_score, "status": i.status.value,
             "created_at": str(i.created_at)}
            for i in inspections
        ],
    }


@router.get("/compliance-trends")
def compliance_trends(clinic_id: Optional[int] = None, days: int = 180,
                      db: Session = Depends(get_db), _=Depends(get_current_user)):
    now = datetime.utcnow()
    since = now - timedelta(days=days)
    q = db.query(Inspection).filter(
        Inspection.submitted_at >= since,
        Inspection.compliance_score.isnot(None),
    )
    if clinic_id:
        q = q.filter(Inspection.clinic_id == clinic_id)
    inspections = q.order_by(Inspection.submitted_at).all()
    return [
        {"date": str(i.submitted_at.date()), "score": i.compliance_score,
         "risk_level": i.risk_level, "clinic_name": i.clinic.name if i.clinic else None}
        for i in inspections
    ]


@router.get("/export/csv")
def export_csv(resource: str = "inspections", db: Session = Depends(get_db),
               _=Depends(require_admin_or_auditor)):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    if resource == "inspections":
        writer.writerow(["ID", "Clinic", "Inspector", "Score", "Risk", "Status", "Submitted"])
        rows = db.query(Inspection).filter(Inspection.compliance_score.isnot(None)).all()
        for r in rows:
            writer.writerow([r.id, r.clinic.name if r.clinic else "", r.inspector.full_name if r.inspector else "",
                              r.compliance_score, r.risk_level, r.status.value if r.status else "", r.submitted_at])
    elif resource == "actions":
        writer.writerow(["ID", "Clinic", "Title", "Status", "Priority", "Due Date", "Assignee"])
        rows = db.query(CorrectiveAction).all()
        for r in rows:
            writer.writerow([r.id, r.clinic.name if r.clinic else "", r.title,
                              r.status.value if r.status else "", r.priority, r.due_date,
                              r.assignee.full_name if r.assignee else ""])
    elif resource == "certifications":
        writer.writerow(["ID", "Name", "Email", "Course", "Score", "Status", "Completed", "Expires"])
        rows = db.query(TeamCertification).all()
        for r in rows:
            writer.writerow([r.id, r.participant_name, r.participant_email,
                              r.course.title if r.course else "", r.score,
                              r.status.value if r.status else "", r.completed_at, r.expires_at])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={resource}.csv"},
    )


@router.get("/export/excel")
def export_excel(resource: str = "inspections", db: Session = Depends(get_db),
                 _=Depends(require_admin_or_auditor)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
    HEADER_ALIGN = Alignment(horizontal="center")

    def write_headers(headers):
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            ws.column_dimensions[get_column_letter(col)].width = 20

    if resource == "inspections":
        ws.title = "Inspections"
        write_headers(["ID", "Clinic", "Clinic Type", "Inspector", "Score", "Risk Level",
                        "Status", "Check-in", "Submitted"])
        for r in db.query(Inspection).order_by(Inspection.created_at.desc()).all():
            ws.append([r.id,
                        r.clinic.name if r.clinic else "",
                        r.clinic.clinic_type.value if r.clinic and r.clinic.clinic_type else "",
                        r.inspector.full_name if r.inspector else "",
                        round(r.compliance_score, 1) if r.compliance_score else "",
                        r.risk_level or "",
                        r.status.value if r.status else "",
                        str(r.checkin_time) if r.checkin_time else "",
                        str(r.submitted_at) if r.submitted_at else ""])

    elif resource == "actions":
        ws.title = "Corrective Actions"
        write_headers(["ID", "Clinic", "Title", "Status", "Priority", "Assigned To",
                        "Due Date", "Resolved At", "Source"])
        for r in db.query(CorrectiveAction).order_by(CorrectiveAction.created_at.desc()).all():
            ws.append([r.id,
                        r.clinic.name if r.clinic else "",
                        r.title,
                        r.status.value if r.status else "",
                        r.priority,
                        r.assignee.full_name if r.assignee else "Unassigned",
                        str(r.due_date) if r.due_date else "",
                        str(r.resolved_at) if r.resolved_at else "",
                        "Manual" if r.is_manual else f"Inspection #{r.inspection_id}"])

    elif resource == "certifications":
        ws.title = "Certifications"
        write_headers(["ID", "Name", "Email", "Course", "Score", "Status",
                        "Attempts", "Completed", "Expires"])
        for r in db.query(TeamCertification).order_by(TeamCertification.created_at.desc()).all():
            ws.append([r.id, r.participant_name, r.participant_email,
                        r.course.title if r.course else "", r.score,
                        r.status.value if r.status else "", r.attempts,
                        str(r.completed_at) if r.completed_at else "",
                        str(r.expires_at) if r.expires_at else ""])

    elif resource == "clinic_scorecard":
        ws.title = "Clinic Scorecard"
        write_headers(["Clinic", "Type", "City", "State", "Avg Score",
                        "Last Score", "Risk Level", "Open Actions", "Last Inspection"])
        for c in db.query(Clinic).filter(Clinic.is_active == True).order_by(Clinic.name).all():
            avg = db.query(func.avg(Inspection.compliance_score)).filter(
                Inspection.clinic_id == c.id, Inspection.compliance_score.isnot(None)).scalar()
            latest = (db.query(Inspection)
                      .filter(Inspection.clinic_id == c.id, Inspection.compliance_score.isnot(None))
                      .order_by(Inspection.submitted_at.desc()).first())
            open_cnt = db.query(CorrectiveAction).filter(
                CorrectiveAction.clinic_id == c.id,
                CorrectiveAction.status.in_([ActionStatus.open, ActionStatus.in_progress])).count()
            ws.append([c.name,
                        c.clinic_type.value if c.clinic_type else "",
                        c.city or "", c.state or "",
                        round(avg, 1) if avg else "",
                        round(latest.compliance_score, 1) if latest else "",
                        latest.risk_level if latest else "",
                        open_cnt,
                        str(latest.submitted_at.date()) if latest and latest.submitted_at else ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={resource}.xlsx"},
    )


# Import needed for excel scorecard
from sqlalchemy import func
